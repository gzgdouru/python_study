#写xlxs文件
import openpyxl
import json
import os

if __name__ == "__main__":
    data = json.load(open("student.txt", "r", encoding="utf8"))

    #create xlsx file
    if not os.path.exists("student.xlsx"):
        wb = openpyxl.Workbook()
        wb.save(filename="student.xlsx")

    # write file
    wb = openpyxl.load_workbook("student.xlsx")
    # sheet = wb.create_sheet()
    sheet = wb.active
    sheet.title = "student"
    for i, key in enumerate(data):
        sheet.cell(row=i + 1, column=1, value=str(key))
        for j, value in enumerate(data[key]):
            sheet.cell(row=i+1, column=j+2, value=str(value))
    wb.save("student.xlsx")
    wb.close()
