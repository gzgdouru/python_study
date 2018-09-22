#写xlxs文件
import openpyxl
import json
import os

if __name__ == "__main__":
    data = json.load(open("city.txt", "r", encoding="utf8"))

    # if not os.path.exists("city.xlsx"):
    #     wb = openpyxl.Workbook()
    #     wb.save("city.xlsx")

    wb = openpyxl.Workbook()
    sheet = wb.active
    for i, key in enumerate(data):
        sheet.cell(row=i+1, column=1, value=key)
        sheet.cell(row=i+1, column=2, value=data[key])
    wb.save("city.xlsx")

